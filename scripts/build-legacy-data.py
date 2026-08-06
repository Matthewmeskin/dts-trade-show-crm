#!/usr/bin/env python3
"""
Build scripts/data/legacy-customers.json from the legacy trade-show master
workbook. See scripts/README-legacy-import.md.

Usage:
    python scripts/build-legacy-data.py DTS_Trade_Show_Master.xlsx scripts/data/legacy-customers.json
"""
import sys
import json
import datetime
from openpyxl import load_workbook


def s(v):
    if v is None:
        return None
    v = str(v).strip()
    return v or None


def num(v):
    if v is None or v == "":
        return None
    try:
        return round(float(v), 2)
    except (TypeError, ValueError):
        return None


def integer(v):
    n = num(v)
    return int(n) if n is not None else None


def norm_status(v):
    v = (v or "").lower()
    if "active" in v:
        return "active"
    if "dormant" in v:
        return "dormant"
    if "not in current tms" in v:
        return "not_in_tms"
    return None


def tier_letter(v):
    if not v:
        return None
    v = v.strip()
    if v.lower().startswith("internal"):
        return "Internal"
    return v[0].upper() if v[0].isalpha() else None


def as_date(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    return s(v)


def main(src, out):
    wb = load_workbook(src, read_only=True, data_only=True)

    m = wb["Master - All Legacy Customers"]
    mrows = list(m.iter_rows(values_only=True))
    H = {name: i for i, name in enumerate(mrows[0])}

    def g(r, name):
        return r[H[name]] if name in H else None

    companies, order = {}, []
    for r in mrows[1:]:
        if not r or not r[0]:
            continue
        name = str(r[0]).strip()
        if not name:
            continue
        companies[name] = {
            "company_name": name,
            "owner_rep": s(g(r, "Current Rep")),
            "sales_status": norm_status(g(r, "Current Status")),
            "priority_tier": tier_letter(g(r, "Priority Tier")),
            "priority_tier_label": s(g(r, "Priority Tier")),
            "website": s(g(r, "Website")),
            "phone": s(g(r, "Phone")),
            "ttm_loads": integer(g(r, "TTM Loads")),
            "ttm_margin": num(g(r, "TTM Margin")),
            "last_pickup": as_date(g(r, "Last Pickup")),
            "legacy_loads": integer(g(r, "Legacy Loads")),
            "legacy_first_year": integer(g(r, "Legacy First Yr")),
            "legacy_last_year": integer(g(r, "Legacy Last Yr")),
            "legacy_billed": num(g(r, "Legacy Billed")),
            "legacy_margin": num(g(r, "Legacy Margin")),
            "legacy_margin_per_load": num(g(r, "Legacy Margin/Load")),
            "shows_shipped": s(g(r, "Shows They Shipped With Us")),
            "shows_confirmed_2026": s(g(r, "2026 Shows Confirmed")),
            "top_show_cities": s(g(r, "Top Show Cities")),
            "history": [],
        }
        order.append(name)

    cs = wb["Customers by Show"]
    crows = list(cs.iter_rows(values_only=True))
    C = {name: i for i, name in enumerate(crows[0])}

    def gc(r, name):
        return r[C[name]] if name in C else None

    for r in crows[1:]:
        if not r:
            continue
        cust = gc(r, "Customer Name")
        if not cust:
            continue
        cust = str(cust).strip()
        if cust not in companies:
            continue
        companies[cust]["history"].append({
            "show_name": s(gc(r, "Show")),
            "show_loads": integer(gc(r, "Show Loads")),
            "first_year": integer(gc(r, "First Yr")),
            "last_year": integer(gc(r, "Last Yr")),
            "billed": num(gc(r, "Show Billed")),
            "margin": num(gc(r, "Show Margin")),
            "confirmed_2026": s(gc(r, "Also Confirmed for 2026?")),
        })

    data = [companies[n] for n in order]
    with open(out, "w") as f:
        json.dump(data, f, ensure_ascii=False, indent=1)
    print(f"Wrote {len(data)} companies, "
          f"{sum(len(c['history']) for c in data)} history rows to {out}")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print(__doc__)
        sys.exit(1)
    main(sys.argv[1], sys.argv[2])
